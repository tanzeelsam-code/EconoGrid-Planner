"""
Excel Formatting Utilities — Shared Excel Export Engine.

Provides professional, publication-quality Excel formatting that
emulates the reporting style of EViews, LEAP, and RETScreen.
Uses openpyxl for full control over cell styles, colors, borders,
merged cells, number formats, and conditional formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from typing import Optional, List, Dict, Any
import pandas as pd
from datetime import datetime
import sys
import os

# Add parent directory to path for config import
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from config import EXCEL_STYLES


class ExcelFormatter:
    """
    Professional Excel workbook formatter.

    Creates publication-quality Excel outputs with consistent styling
    across all three analytical modules.
    """

    def __init__(self, title: str = "Analysis Output"):
        """
        Initialize formatter with a new workbook.

        Args:
            title: Workbook title for the properties metadata.
        """
        self.wb = Workbook()
        self.wb.properties.title = title
        self.wb.properties.creator = "EconoGrid Planner"
        self.wb.properties.created = datetime.now()

        # Remove default sheet — we'll create named sheets
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Pre-build style objects
        self._header_fill = PatternFill(
            start_color=EXCEL_STYLES["header_fill"],
            end_color=EXCEL_STYLES["header_fill"],
            fill_type="solid"
        )
        self._subheader_fill = PatternFill(
            start_color=EXCEL_STYLES["subheader_fill"],
            end_color=EXCEL_STYLES["subheader_fill"],
            fill_type="solid"
        )
        self._accent_fill = PatternFill(
            start_color=EXCEL_STYLES["accent_fill"],
            end_color=EXCEL_STYLES["accent_fill"],
            fill_type="solid"
        )
        self._positive_fill = PatternFill(
            start_color=EXCEL_STYLES["positive_fill"],
            end_color=EXCEL_STYLES["positive_fill"],
            fill_type="solid"
        )
        self._negative_fill = PatternFill(
            start_color=EXCEL_STYLES["negative_fill"],
            end_color=EXCEL_STYLES["negative_fill"],
            fill_type="solid"
        )
        self._header_font = Font(
            name="Calibri",
            size=EXCEL_STYLES["header_font_size"],
            bold=True,
            color=EXCEL_STYLES["header_font_color"]
        )
        self._title_font = Font(
            name="Calibri",
            size=EXCEL_STYLES["title_font_size"],
            bold=True,
            color=EXCEL_STYLES["header_fill"]
        )
        self._body_font = Font(
            name="Calibri",
            size=EXCEL_STYLES["body_font_size"]
        )
        self._bold_font = Font(
            name="Calibri",
            size=EXCEL_STYLES["body_font_size"],
            bold=True
        )
        self._thin_border = Border(
            left=Side(style="thin", color=EXCEL_STYLES["border_color"]),
            right=Side(style="thin", color=EXCEL_STYLES["border_color"]),
            top=Side(style="thin", color=EXCEL_STYLES["border_color"]),
            bottom=Side(style="thin", color=EXCEL_STYLES["border_color"])
        )
        self._center_align = Alignment(horizontal="center", vertical="center")
        self._left_align = Alignment(horizontal="left", vertical="center")
        self._right_align = Alignment(horizontal="right", vertical="center")
        self._wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def add_sheet(self, name: str) -> Any:
        """Add a new worksheet with the given name."""
        ws = self.wb.create_sheet(title=name[:31])  # Excel max 31 chars
        return ws

    def write_title_block(
        self,
        ws,
        title: str,
        subtitle: str = "",
        row: int = 1,
        col: int = 1,
        span: int = 6
    ) -> int:
        """
        Write a title block at the top of a sheet.

        Args:
            ws: Worksheet object.
            title: Main title string.
            subtitle: Optional subtitle / description.
            row: Starting row (1-indexed).
            col: Starting column (1-indexed).
            span: Number of columns the title should span.

        Returns:
            Next available row after the title block.
        """
        # Title
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = self._title_font
        cell.alignment = self._left_align
        if span > 1:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=col + span - 1
            )

        # Subtitle
        if subtitle:
            row += 1
            cell = ws.cell(row=row, column=col, value=subtitle)
            cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
            cell.alignment = self._left_align
            if span > 1:
                ws.merge_cells(
                    start_row=row, start_column=col,
                    end_row=row, end_column=col + span - 1
                )

        # Timestamp
        row += 1
        cell = ws.cell(
            row=row, column=col,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        cell.font = Font(name="Calibri", size=9, italic=True, color="999999")

        return row + 2  # Leave a blank row

    def write_header_row(
        self,
        ws,
        headers: List[str],
        row: int,
        col: int = 1,
        style: str = "primary"
    ) -> None:
        """
        Write a styled header row.

        Args:
            ws: Worksheet object.
            headers: List of header strings.
            row: Row number (1-indexed).
            col: Starting column (1-indexed).
            style: "primary" for dark fill, "secondary" for medium fill.
        """
        fill = self._header_fill if style == "primary" else self._subheader_fill
        for i, header in enumerate(headers):
            cell = ws.cell(row=row, column=col + i, value=header)
            cell.font = self._header_font
            cell.fill = fill
            cell.alignment = self._center_align
            cell.border = self._thin_border

    def write_dataframe(
        self,
        ws,
        df: pd.DataFrame,
        start_row: int,
        start_col: int = 1,
        include_header: bool = True,
        include_index: bool = True,
        number_format: Optional[str] = None,
        header_style: str = "primary"
    ) -> int:
        """
        Write a pandas DataFrame to Excel with professional formatting.

        Args:
            ws: Worksheet object.
            df: DataFrame to write.
            start_row: Starting row (1-indexed).
            start_col: Starting column (1-indexed).
            include_header: Whether to write column headers.
            include_index: Whether to write the index.
            number_format: Optional format string for numeric cells.
            header_style: "primary" or "secondary".

        Returns:
            Next available row after the data.
        """
        current_row = start_row
        col_offset = 1 if include_index else 0

        # Write headers
        if include_header:
            headers = []
            if include_index:
                idx_name = df.index.name if df.index.name else ""
                headers.append(idx_name)
            headers.extend(df.columns.tolist())
            self.write_header_row(ws, headers, current_row, start_col, header_style)
            current_row += 1

        # Write data rows
        for idx, (row_label, row_data) in enumerate(df.iterrows()):
            # Index column
            if include_index:
                cell = ws.cell(row=current_row, column=start_col, value=row_label)
                cell.font = self._bold_font
                cell.alignment = self._left_align
                cell.border = self._thin_border
                # Alternate row shading
                if idx % 2 == 0:
                    cell.fill = self._accent_fill

            # Data columns
            for j, value in enumerate(row_data):
                cell = ws.cell(
                    row=current_row,
                    column=start_col + col_offset + j,
                    value=value
                )
                cell.font = self._body_font
                cell.alignment = self._right_align
                cell.border = self._thin_border

                # Apply number format
                if number_format and isinstance(value, (int, float)):
                    cell.number_format = number_format

                # Alternate row shading
                if idx % 2 == 0:
                    cell.fill = self._accent_fill

            current_row += 1

        return current_row + 1  # Leave one blank row

    def write_key_value_block(
        self,
        ws,
        data: Dict[str, Any],
        start_row: int,
        start_col: int = 1,
        title: Optional[str] = None,
        number_format: Optional[str] = None
    ) -> int:
        """
        Write a key-value statistics block (like EViews model stats).

        Args:
            ws: Worksheet object.
            data: Dictionary of {label: value} pairs.
            start_row: Starting row.
            start_col: Starting column.
            title: Optional block title.
            number_format: Format for numeric values.

        Returns:
            Next available row.
        """
        current_row = start_row

        if title:
            cell = ws.cell(row=current_row, column=start_col, value=title)
            cell.font = self._bold_font
            cell.fill = self._subheader_fill
            cell.font = Font(
                name="Calibri", size=11, bold=True,
                color=EXCEL_STYLES["header_font_color"]
            )
            ws.cell(row=current_row, column=start_col + 1).fill = self._subheader_fill
            ws.cell(row=current_row, column=start_col + 1).border = self._thin_border
            cell.border = self._thin_border
            current_row += 1

        for i, (key, value) in enumerate(data.items()):
            # Label
            cell = ws.cell(row=current_row, column=start_col, value=key)
            cell.font = self._body_font
            cell.alignment = self._left_align
            cell.border = self._thin_border
            if i % 2 == 0:
                cell.fill = self._accent_fill

            # Value
            cell = ws.cell(row=current_row, column=start_col + 1, value=value)
            cell.font = self._body_font
            cell.alignment = self._right_align
            cell.border = self._thin_border
            if number_format and isinstance(value, (int, float)):
                cell.number_format = number_format
            if i % 2 == 0:
                cell.fill = self._accent_fill

            current_row += 1

        return current_row + 1

    def auto_fit_columns(self, ws, min_width: int = 10, max_width: int = 35) -> None:
        """Auto-fit column widths based on content."""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, length)
                except:
                    pass
            adjusted_width = min(max(max_length + 3, min_width), max_width)
            ws.column_dimensions[col_letter].width = adjusted_width

    def apply_conditional_fill(
        self,
        ws,
        row: int,
        col: int,
        value: float,
        positive_good: bool = True
    ) -> None:
        """Apply green/red fill based on positive/negative value."""
        cell = ws.cell(row=row, column=col)
        if value > 0:
            cell.fill = self._positive_fill if positive_good else self._negative_fill
        elif value < 0:
            cell.fill = self._negative_fill if positive_good else self._positive_fill

    def save(self, filepath: str) -> str:
        """
        Save workbook to file.

        Args:
            filepath: Output file path.

        Returns:
            Absolute path of saved file.
        """
        self.wb.save(filepath)
        return os.path.abspath(filepath)
